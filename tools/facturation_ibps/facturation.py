# -*- coding: utf-8 -*-


import argparse
import re
import warnings

import openpyxl

import pandas as pd


warnings.filterwarnings("ignore")


def Parser():
    the_parser = argparse.ArgumentParser()
    the_parser.add_argument('--input', '-i', action='store', type=str,
                            help="input html code to convert to xlsx")
    the_parser.add_argument('--output', '-o', action='store', type=str,
                            help='xlsx converted file')
    the_parser.add_argument('--template', '-t', action='store', type=str,
                            help='xlsx template file')
    the_parser.add_argument('--reduction', '-r', action='store', type=float,
                            help='reduction to apply', default=1.0)
    args = the_parser.parse_args()
    return args


def main(template, input_file, output_file, reduction):
    """Script de parsing des fichiers de facturation de l'IBPS"""

    # ouverture fichier input
    with open(input_file, 'rb') as file_object:
        facture_html = file_object.read()
    # convert to unicode utf-8, remove &nbsp and €
    facture_html = facture_html.decode('utf-8')
    facture_html = facture_html.replace(r'&nbsp;', r' ')
    facture_html = facture_html.replace(r' &euro;', '')
    facture_html = facture_html.replace(u' \u20ac', '')
    # parsing de la référence, de la date et de la période de facturation
    date = re.search(r'Paris le (.*?)</p>',
                     facture_html).group(1)
    periode = re.search(r'de la prestation (.*?)</p>',
                        facture_html).group(1)
    ref = re.search(r'rence interne d.*? :\s*(.*?)<',
                    facture_html).group(1)

    # parsing des tableaux html avec pandas
    facture_parsed = pd.read_html(
        facture_html,
        thousands='',
        decimal='.',
        flavor='bs4')
    # remove 'Adresse de l'appel à facturation : ' (\xa0:\xa0)
    adresse = facture_parsed[0].replace(
        r"Adresse de l\'appel \xe0 facturation : ", r'', regex=True)
    adresse = adresse.replace(
        r"Adresse du client : ", r'', regex=True)
    elements = facture_parsed[1]

    # conversion des noms de colonnes
    elements_col = elements.iloc[0]
    cout_col = elements_col.str.extract(r'(cout.*)',
                                        expand=False).dropna().iloc[0]
    elements = elements.rename(columns=elements_col).drop(
        elements.index[0])

    # changement du type des éléments numériques du tableau
    elements[u'nombre(s)'] = pd.to_numeric(elements[u'nombre(s)'])
    elements[cout_col] = pd.to_numeric(elements[cout_col])

    # ouverture fichier output
    facture_output = openpyxl.load_workbook(
        template, data_only=False, keep_vba=False)
    ws = facture_output.worksheets[0]

    # rajout de l'image de SU qui ne survit pas à la conversion
    img = openpyxl.drawing.image.Image('template_SU.jpg')
    img.anchor = "A1"
    ws.add_image(img)

    # ajout des éléments facturés dans le tableau
    element_row = 23
    for i in range(len(elements)):
        element_row += 1
        ws.cell(row=element_row, column=1, value=elements.iloc[i][u'Objet'])
        ws.cell(
            row=element_row,
            column=2,
            value=elements.iloc[i][u'nombre(s)']).number_format = '0.00'
        ws.cell(
            row=element_row,
            column=4,
            value=((1-reduction) *
                   elements.iloc[i][cout_col])).number_format = '0.00'

    # ajout de l'adresse
    address_row = 7
    for i in range(len(adresse)):
        address_row += 1
        ws.cell(row=address_row, column=3,
                value=adresse.iloc[i, 0].encode('utf-8'))

    # ajout de la référence/période/date
    ws.cell(row=2, column=3, value=ref.encode('utf-8'))
    ws.cell(row=5, column=5, value=periode.encode('utf-8'))
    ws.cell(row=21, column=5, value=date.encode('utf-8'))

    # export fichier output
    facture_output.save(output_file)
    return


if __name__ == '__main__':
    args = Parser()
    main(args.template, args.input, args.output, args.reduction)
